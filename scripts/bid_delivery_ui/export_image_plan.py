#!/usr/bin/env python3
"""Export the one confirmed image-planning workbook on any supported host.

``@oai/artifact-tool`` is used when supplied by a host, but it is never a
hard dependency: the portable openpyxl renderer produces the same hand-off
workbook when an AI host does not ship OpenAI-internal Node packages.
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

try:
    from . import protocol
    from .image_prompt import prompt_for_image
except ImportError:
    import protocol
    from image_prompt import prompt_for_image


SCRIPT = Path(__file__).with_name("export_image_plan.mjs")
HEADERS = ["图号", "一级章节", "精确放置位置", "具体放置说明", "图片名称", "图片类型", "用途/核心表达", "核心节点", "构图建议", "画面方向", "是否章首总览图", "统一视觉要求", "应避免的风格", "生图补充说明", "AI生图提示词"]


def _optional_runtime(argument: str | None, environment_name: str) -> Path | None:
    raw = argument or os.environ.get(environment_name)
    if not raw:
        return None
    path = Path(raw).expanduser().resolve()
    return path if path.exists() else None


def _visual_values(visual: dict[str, Any]) -> tuple[str, str]:
    def value(english: str, chinese: str) -> Any:
        return visual.get(english) or visual.get(chinese)
    visual_text = "；".join(str(item) for item in (value("palette", "主色"), value("style", "风格"), value("background", "背景"), value("density", "信息密度")) if item)
    avoid = value("avoid", "应避免")
    avoid_text = "、".join(str(item) for item in avoid) if isinstance(avoid, list) else str(avoid or "")
    return visual_text, avoid_text


def _source_rows(source: dict[str, Any]) -> list[list[str]]:
    visual_text, avoid_text = _visual_values(source.get("visual_direction") or {})
    orientations = {"landscape": "横向", "portrait": "纵向", "square": "方形", "auto": "自适应"}
    rows: list[list[str]] = []
    for image in source["images"]:
        position = image["position"]
        rows.append([
            image["figure_no"], f"{image['chapter_number']} {image['chapter_title']}",
            f"{position['outline_number']} {position['outline_title']}", position["placement_note"],
            image["name"], image["type"], image["purpose"], "；".join(image.get("core_nodes") or []),
            image["composition"], orientations.get(image["orientation"], str(image["orientation"])),
            "是" if image["is_chapter_overview"] else "否", visual_text, avoid_text,
            "本表仅定义图片内容、构图要求和逐图生图提示词；最终交付后仅在用户明确回复“继续”时进入可选生图流程。",
            prompt_for_image(image, source.get("visual_direction") or {}),
        ])
    return rows


def _export_with_openpyxl(source_path: Path, output_path: Path) -> dict[str, Any]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise ValueError("当前环境既无可用的图片规划渲染器，也缺少 openpyxl；请安装 openpyxl 后重试") from exc
    source = json.loads(source_path.read_text(encoding="utf-8"))
    rows = _source_rows(source)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "图片规划清单"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:O1")
    sheet["A1"] = "图片规划表（含逐图AI生图提示词）"
    sheet["A1"].font = Font(name="仿宋_GB2312", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="C91F37")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells("A2:O2")
    sheet["A2"] = "本表包含逐图 AI 生图提示词；最终交付后只有用户明确回复“继续”才进入可选本机生图流程，默认不生成图片。"
    sheet["A2"].font = Font(name="仿宋_GB2312", italic=True, color="7A4B00")
    sheet["A2"].fill = PatternFill("solid", fgColor="FFF5E6")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 34
    for column, header in enumerate(HEADERS, 1):
        cell = sheet.cell(4, column, header)
        cell.font = Font(name="仿宋_GB2312", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E5EAA")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [12, 20, 24, 30, 22, 16, 28, 30, 36, 12, 16, 36, 24, 38, 72]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    border = Border(*(Side(style="thin", color="D9E0EA") for _ in range(4)))
    for row_index, values in enumerate(rows, 5):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="仿宋_GB2312", size=11)
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="center")
            cell.border = border
        sheet.row_dimensions[row_index].height = 54
    for cell in sheet[4]:
        cell.border = border
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:O{max(4, 4 + len(rows))}"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    checked = load_workbook(output_path, read_only=True, data_only=True)
    return {"row_count": max(checked.active.max_row - 4, 0), "worksheet_count": len(checked.worksheets), "renderer": "openpyxl-fallback"}


def _export_with_artifact_tool(source_path: Path, output_path: Path, preview_path: Path | None, node_path: Path, modules_path: Path) -> dict[str, Any]:
    with tempfile.TemporaryDirectory(prefix="biaoshu-image-plan-") as temp_dir:
        temp = Path(temp_dir)
        runtime_script = temp / SCRIPT.name
        shutil.copy2(SCRIPT, runtime_script)
        (temp / "node_modules").symlink_to(modules_path, target_is_directory=True)
        result_path = temp / "renderer-result.json"
        command = [str(node_path), str(runtime_script), str(source_path), str(output_path), str(preview_path) if preview_path else "", str(result_path)]
        run = subprocess.run(command, capture_output=True, text=True, check=False)
        result_text = result_path.read_text(encoding="utf-8") if result_path.exists() else ""
    if run.returncode:
        detail = (run.stderr or run.stdout).strip()[-1000:]
        raise ValueError(detail or "artifact-tool 未返回成功结果")
    result = json.loads(result_text)
    result["renderer"] = "artifact-tool"
    return result


def export_image_plan(project_dir: Path, *, node: str | None = None, node_modules: str | None = None, render_preview: bool = False) -> dict[str, Any]:
    manifest = protocol.load_manifest(project_dir)
    workbook = manifest["image_plan_workbook"]
    source_path = protocol.delivery_dir(project_dir) / workbook["source_path"]
    if workbook["status"] == "pending" or not source_path.exists():
        protocol.build_image_plan_source(project_dir)
        manifest = protocol.load_manifest(project_dir)
        workbook = manifest["image_plan_workbook"]
    if workbook["status"] not in {"generating", "export_pending"}:
        raise ValueError("图片规划Excel当前不处于可导出状态")
    output_path = protocol.delivery_dir(project_dir) / workbook["export_path"]
    preview_path = protocol.delivery_dir(project_dir) / protocol.RESULTS_DIR_NAME / "image-plan-preview.png" if render_preview else None
    node_path, modules_path = _optional_runtime(node, "BIAOSHU_NODE"), _optional_runtime(node_modules, "BIAOSHU_NODE_MODULES")
    fallback_reason = None
    use_artifact_tool = os.environ.get("BIAOSHU_USE_ARTIFACT_TOOL") == "1"
    if use_artifact_tool and node_path and modules_path:
        try:
            if preview_path:
                preview_path.parent.mkdir(parents=True, exist_ok=True)
            renderer_result = _export_with_artifact_tool(source_path, output_path, preview_path, node_path, modules_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            fallback_reason = str(exc)
            renderer_result = _export_with_openpyxl(source_path, output_path)
    else:
        fallback_reason = "使用跨宿主 openpyxl 基础导出器"
        renderer_result = _export_with_openpyxl(source_path, output_path)
    if not output_path.is_file() or output_path.stat().st_size < 512:
        raise ValueError("图片规划Excel导出文件无效")
    validation = {"schema_version": 1, "kind": "bid_delivery_image_plan_validation", "project_id": manifest["project_id"], "checked_at": protocol.utc_now(), "source_sha256": protocol.sha256_file(source_path), "export_sha256": protocol.sha256_file(output_path), "checks": {"xlsx_exists": True, "worksheet_count": renderer_result.get("worksheet_count"), "image_record_count": renderer_result.get("row_count"), "renderer": renderer_result.get("renderer"), "artifact_tool_fallback_reason": fallback_reason, "image_generation_in_scope": False, "image_insertion_in_scope": False}}
    protocol.atomic_write_json(protocol.delivery_dir(project_dir) / protocol.RESULTS_DIR_NAME / "image-plan-validation.json", validation)
    updated = protocol.register_image_plan_artifacts(project_dir)
    return {"manifest": updated, "output": str(output_path), "validation": validation, "preview": str(preview_path) if preview_path and preview_path.exists() else None}


def main() -> int:
    parser = argparse.ArgumentParser(description="导出主标图片规划Excel")
    parser.add_argument("project_dir", type=Path); parser.add_argument("--node"); parser.add_argument("--node-modules"); parser.add_argument("--render-preview", action="store_true")
    args = parser.parse_args()
    result = export_image_plan(args.project_dir.expanduser().resolve(), node=args.node, node_modules=args.node_modules, render_preview=args.render_preview)
    print(json.dumps({"output": result["output"], "status": result["manifest"]["image_plan_workbook"]["status"], "validation": result["validation"], "preview": result["preview"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
